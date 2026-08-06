from __future__ import annotations

import base64
import io
import unicodedata
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class MappingRule:
    ordem: int
    especie: str
    estratificacao: str
    sexo: str
    estado: str
    classificacao: str
    chave: str
    ativo: bool


def _texto(value) -> str:
    return str(value or '').strip()


def _normalizar(value: str) -> str:
    text = unicodedata.normalize('NFKD', _texto(value))
    return ''.join(c for c in text if not unicodedata.combining(c)).upper()


def _bytes_xlsm(source=None) -> bytes:
    if source is None:
        source = Path(__file__).resolve().parents[2] / 'static' / 'classificacao_rebanho_fichas.xlsm'
    if hasattr(source, '__fspath__'):
        raw = Path(source).read_bytes()
    elif isinstance(source, bytes):
        raw = source
    else:
        raw = source.read()

    if raw[:2] != b'PK':
        try:
            decoded = base64.b64decode(raw, validate=True)
        except Exception as exc:
            raise ValueError('Fonte do MAPEAMENTO não é um XLSM válido.') from exc
        if decoded[:2] != b'PK':
            raise ValueError('Fonte do MAPEAMENTO não contém um XLSM válido.')
        raw = decoded
    return raw


class MappingCatalog:
    def __init__(self, rules: list[MappingRule]):
        self.rules = tuple(rules)
        self._by_key = {
            self._key(r.estado, r.sexo, r.estratificacao): r
            for r in self.rules if r.ativo
        }

    @staticmethod
    def _key(estado, sexo, estratificacao) -> str:
        return '|'.join((
            _normalizar(estado),
            _normalizar(sexo),
            _normalizar(estratificacao),
        ))

    def lookup(self, estado, sexo, estratificacao) -> MappingRule | None:
        estado = {
            'GO_DEC_WEB': 'GO_DECLARACAO',
            'AGRODEFESA_GO': 'GO_DECLARACAO',
        }.get(_normalizar(estado), estado)
        return self._by_key.get(self._key(estado, sexo, estratificacao))

    def __len__(self):
        return len(self.rules)


def load_mapping(source=None) -> MappingCatalog:
    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(_bytes_xlsm(source)),
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    if 'MAPEAMENTO' not in workbook.sheetnames:
        raise ValueError('Aba MAPEAMENTO não encontrada no XLSM.')

    ws = workbook['MAPEAMENTO']
    rows = ws.iter_rows(values_only=True)
    header = next((row for row in rows if any(value is not None for value in row)), None)
    if not header:
        raise ValueError('Aba MAPEAMENTO vazia.')
    columns = {_normalizar(value): index for index, value in enumerate(header)}
    required = {
        'ORDEM', 'ESPECIE', 'ESTRATIFICACAO', 'SEXO',
        'ESTADO', 'CLASSIFICACAO', 'CHAVE', 'ATIVO',
    }
    missing = required - columns.keys()
    if missing:
        raise ValueError(f'Colunas ausentes no MAPEAMENTO: {sorted(missing)}')

    rules = []
    for row in rows:
        values = list(row)
        get = lambda name: values[columns[name]] if columns[name] < len(values) else ''
        ativo = _normalizar(get('ATIVO')) in {'SIM', 'S', 'TRUE', '1'}
        rules.append(MappingRule(
            ordem=int(get('ORDEM') or 0),
            especie=_texto(get('ESPECIE')).upper(),
            estratificacao=_texto(get('ESTRATIFICACAO')).upper(),
            sexo=_texto(get('SEXO')).upper(),
            estado=_texto(get('ESTADO')).upper(),
            classificacao=_texto(get('CLASSIFICACAO')),
            chave=_texto(get('CHAVE')),
            ativo=ativo,
        ))
    return MappingCatalog(rules)
