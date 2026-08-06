"""Fonte única das regras da aba MAPEAMENTO do XLSM de fichas."""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import unicodedata


_ARQUIVO_PADRAO = Path(__file__).resolve().parents[1] / 'static' / 'ficha_consolidado_exemplo.xlsm'
_COLUNAS = ('ORDEM', 'ESPÉCIE', 'ESTRATIFICAÇÃO', 'SEXO', 'ESTADO', 'CLASSIFICAÇÃO', 'CHAVE', 'ATIVO')


def _normalizar(valor) -> str:
    texto = str(valor or '').strip().upper()
    return ''.join(
        c for c in unicodedata.normalize('NFKD', texto)
        if not unicodedata.combining(c)
    )


def _estado_mapeamento(estado: str) -> str:
    origem = _normalizar(estado).replace('-', '_').replace(' ', '_')
    return {
        'MT': 'MT_DECLARACAO',
        'RO': 'RO_DECLARACAO',
        'PA': 'PA_DECLARACAO',
        'TO': 'TO_DECLARACAO',
        'GO': 'GO_DECLARACAO',
        'GO_DEC_WEB': 'GO_DECLARACAO',
        'AGRODEFESA_GO': 'GO_DECLARACAO',
        'GO_IR': 'GO IR',
    }.get(origem, origem)


@lru_cache(maxsize=4)
def load_mapeamento(source: str | Path | None = None) -> dict[str, dict]:
    """Carrega as regras ativas da aba MAPEAMENTO do XLSM do projeto."""
    import openpyxl

    caminho = Path(source) if source else _ARQUIVO_PADRAO
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True, keep_vba=True)
    if 'MAPEAMENTO' not in wb.sheetnames:
        raise ValueError('Aba MAPEAMENTO não encontrada no XLSM')
    ws = wb['MAPEAMENTO']

    cabecalho = None
    for linha in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        normalizada = tuple(_normalizar(c) for c in linha[:len(_COLUNAS)])
        if normalizada == tuple(_normalizar(c) for c in _COLUNAS):
            cabecalho = linha
            break
    if cabecalho is None:
        raise ValueError('Cabeçalho esperado não encontrado na aba MAPEAMENTO')

    regras = {}
    for linha in ws.iter_rows(min_row=ws.min_row + 1, values_only=True):
        valores = list(linha[:len(_COLUNAS)])
        if len(valores) < len(_COLUNAS) or _normalizar(valores[7]) != 'SIM':
            continue
        regra = dict(zip(_COLUNAS, valores))
        chave = str(regra['CHAVE']).strip()
        if chave in {'', 'None'}:
            chave = f"{regra['ESTADO']}|{regra['SEXO']}|{regra['ESTRATIFICAÇÃO']}"
        regras[chave] = {
            'ordem': int(regra['ORDEM']),
            'especie': str(regra['ESPÉCIE']).strip(),
            'estratificacao': str(regra['ESTRATIFICAÇÃO']).strip(),
            'sexo': str(regra['SEXO']).strip(),
            'estado': str(regra['ESTADO']).strip(),
            'classificacao': str(regra['CLASSIFICAÇÃO']).strip(),
            'chave': chave,
            'ativo': str(regra['ATIVO']).strip(),
        }
    return regras


def buscar_mapeamento(estado: str, sexo: str, estratificacao: str) -> dict:
    """Busca uma regra por estado, sexo e estratificação, como a macro."""
    estado_excel = _estado_mapeamento(estado)
    sexo_excel = _normalizar(sexo)
    faixa_excel = _normalizar(estratificacao)
    regras = load_mapeamento()
    for regra in regras.values():
        if (
            _normalizar(regra['estado']) == _normalizar(estado_excel)
            and _normalizar(regra['sexo']) == sexo_excel
            and _normalizar(regra['estratificacao']) == faixa_excel
        ):
            return regra
    raise KeyError(f'Mapeamento não encontrado: {estado}|{sexo}|{estratificacao}')


def mapear_animais(animais: dict, estado: str) -> list[dict]:
    """Distribui o rebanho nas classificações definidas no XLSM."""
    resultado = []
    estado_excel = _estado_mapeamento(estado)
    origem = _normalizar(estado).replace('-', '_').replace(' ', '_')
    # GO/MT possuem regras de 0-4 e 5-12 no modelo INDEA. As linhas
    # agregadas de 0-12 tambem existem na planilha para outros modelos,
    # mas nao podem ser somadas junto com as faixas divididas.
    faixas_divididas = origem in {'MT', 'GO', 'GO_DEC_WEB', 'GO_IR'}
    regras = load_mapeamento()
    for regra in regras.values():
        if _normalizar(regra['estado']) != _normalizar(estado_excel):
            continue
        sexo_key = 'F' if _normalizar(regra['sexo']) == 'FEMEA' else 'M'
        faixa = _normalizar(regra['estratificacao'])
        if faixas_divididas and faixa == '0 A 12 MESES':
            continue
        if faixa == '0 A 12 MESES':
            quantidade = animais.get(f'f00_{sexo_key}', 0) + animais.get(f'f05_{sexo_key}', 0)
        elif faixa == '00 A 04 MESES':
            quantidade = animais.get(f'f00_{sexo_key}', 0)
        elif faixa == '05 A 12 MESES':
            quantidade = animais.get(f'f05_{sexo_key}', 0)
        elif faixa == '13 A 24 MESES':
            quantidade = animais.get(f'f13_{sexo_key}', 0)
        elif faixa == '25 A 36 MESES':
            quantidade = animais.get(f'f25_{sexo_key}', 0)
        elif faixa == 'ACIMA DE 36 MESES':
            quantidade = animais.get(f'fac_{sexo_key}', 0)
        else:
            quantidade = 0
        if quantidade:
            resultado.append({**regra, 'quantidade': int(quantidade)})
    return resultado
