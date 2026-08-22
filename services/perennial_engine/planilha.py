"""Leitura da ficha de talhões (.xlsx) para o payload da análise perene.

Lê por NOME DE COLUNA, não por posição: o analista que insere uma coluna no
meio da planilha não deve mudar silenciosamente o significado dos números.

Devolve o mesmo payload que /api/perene/analisar recebe, mais a lista do que
está faltando. Falta não é exceção — é resultado: a ficha meio preenchida é o
caso comum, e o analista precisa saber o que completar, não receber um erro.
"""
from __future__ import annotations

import io
import unicodedata
from typing import Any

_FASES = {'alta', 'baixa'}


def _normalizar(valor) -> str:
    texto = unicodedata.normalize('NFKD', str(valor or '').strip())
    return ''.join(c for c in texto if not unicodedata.combining(c)).upper()


def _float(valor):
    if valor is None or isinstance(valor, bool):
        return None
    try:
        return float(str(valor).replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _int(valor):
    numero = _float(valor)
    return int(numero) if numero is not None else None


def _linhas(ws) -> list[dict[str, Any]]:
    """Cabeçalho na primeira linha; devolve uma linha por dicionário."""
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []
    colunas = [_normalizar(c) for c in linhas[0]]
    saida = []
    for valores in linhas[1:]:
        registro = {
            coluna: valores[indice] if indice < len(valores) else None
            for indice, coluna in enumerate(colunas) if coluna
        }
        if any(v is not None and str(v).strip() != '' for v in registro.values()):
            saida.append(registro)
    return saida


def _talhoes(ws, faltando: list[str]) -> list[dict[str, Any]]:
    talhoes = []
    for numero, linha in enumerate(_linhas(ws), start=2):
        cultura = _normalizar(linha.get('CULTURA'))
        area = _float(linha.get('AREA (HA)'))
        plantio = _int(linha.get('ANO DE PLANTIO'))
        if not cultura:
            faltando.append(f'TALHOES linha {numero}: cultura em branco.')
            continue
        if not area or area <= 0:
            faltando.append(f'TALHOES linha {numero}: área ausente ou inválida.')
            continue
        if not plantio:
            faltando.append(f'TALHOES linha {numero}: ano de plantio ausente.')
            continue
        fase = str(linha.get('FASE DE CARGA') or '').strip().lower() or None
        if fase and fase not in _FASES:
            faltando.append(
                f'TALHOES linha {numero}: fase de carga deve ser alta ou baixa.')
            fase = None
        talhoes.append({
            'cultura': cultura,
            'area_ha': area,
            'ano_plantio': plantio,
            'identificacao': str(linha.get('IDENTIFICACAO') or '').strip(),
            'fase_bienal': fase,
        })
    return talhoes


def _curvas(ws, faltando: list[str]) -> dict[str, dict[str, Any]]:
    curvas: dict[str, dict[str, Any]] = {}
    for numero, linha in enumerate(_linhas(ws), start=2):
        cultura = _normalizar(linha.get('CULTURA'))
        if not cultura:
            continue
        plena = _float(linha.get('PRODUTIVIDADE PLENA'))
        if not plena or plena <= 0:
            faltando.append(
                f'CURVAS linha {numero}: produtividade plena de {cultura} ausente.')
            continue
        fatores = {}
        for coluna, valor in linha.items():
            if not coluna.startswith('IDADE '):
                continue
            fator = _float(valor)
            if fator is not None:
                fatores[int(coluna.split(' ')[1])] = fator
        if not fatores:
            faltando.append(
                f'CURVAS linha {numero}: nenhum fator por idade para {cultura}.')
            continue
        curvas[cultura] = {
            'produtividade_plena': plena,
            'unidade': str(linha.get('UNIDADE') or '').strip() or 'unidade',
            'fatores': fatores,
            'bienalidade': _float(linha.get('BIENALIDADE')) or 0.0,
            'ciclo_anos': _int(linha.get('CICLO (ANOS)')),
            'fonte': str(linha.get('FONTE') or '').strip(),
        }
    return curvas


def _precos_e_custos(ws) -> tuple[dict[str, float], dict[str, dict[str, float]]]:
    precos: dict[str, float] = {}
    custos: dict[str, dict[str, float]] = {}
    for linha in _linhas(ws):
        cultura = _normalizar(linha.get('CULTURA'))
        if not cultura:
            continue
        preco = _float(linha.get('PRECO POR UNIDADE'))
        if preco is not None:
            precos[cultura] = preco
        tabela = {
            'formacao': _float(linha.get('CUSTO FORMACAO (R$/HA)')),
            'producao': _float(linha.get('CUSTO PRODUCAO (R$/HA)')),
            'reforma': _float(linha.get('CUSTO REFORMA (R$/HA)')),
            'por_unidade': _float(linha.get('CUSTO COLHEITA POR UNIDADE')),
        }
        tabela = {k: v for k, v in tabela.items() if v is not None}
        if tabela:
            custos[cultura] = tabela
    return precos, custos


def parsear_ficha_talhoes(source, ano_base: int | None = None) -> dict[str, Any]:
    """Lê a ficha e devolve o payload da análise, com o que falta declarado."""
    import openpyxl

    if isinstance(source, bytes):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)

    faltando: list[str] = []
    for aba in ('TALHOES', 'CURVAS', 'PRECOS'):
        if aba not in wb.sheetnames:
            faltando.append(f'A ficha não tem a aba {aba}.')

    talhoes = _talhoes(wb['TALHOES'], faltando) if 'TALHOES' in wb.sheetnames else []
    curvas = _curvas(wb['CURVAS'], faltando) if 'CURVAS' in wb.sheetnames else {}
    precos, custos = (
        _precos_e_custos(wb['PRECOS']) if 'PRECOS' in wb.sheetnames else ({}, {}))

    if not talhoes:
        faltando.append('Nenhum talhão preenchido na aba TALHOES.')

    culturas = {t['cultura'] for t in talhoes}
    for cultura in sorted(culturas - set(curvas)):
        faltando.append(f'Sem curva de produtividade para {cultura}.')
    for cultura in sorted(culturas - set(precos)):
        faltando.append(f'Sem preço declarado para {cultura}.')

    sem_fonte = sorted(c for c, dados in curvas.items() if not dados['fonte'])
    avisos = []
    if sem_fonte:
        avisos.append(
            'Curva sem fonte declarada para: ' + ', '.join(sem_fonte)
            + '. Entra no parecer como declaração do analista.')

    payload = {
        'ano_base': ano_base,
        'talhoes': talhoes,
        'curvas': curvas,
        'precos': precos,
        'custos': custos,
    }
    return {
        'completo': not faltando,
        'payload': payload,
        'faltando': faltando,
        'avisos': avisos,
        'culturas': tuple(sorted(culturas)),
        'area_total_ha': round(sum(t['area_ha'] for t in talhoes), 4),
    }
