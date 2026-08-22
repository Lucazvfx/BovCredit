"""Gera a ficha de talhões de lavoura perene (.xlsx, sem macros).

Três abas, todas tabela simples com cabeçalho na primeira linha — o parser em
services/perennial_engine/planilha.py lê por nome de coluna, não por posição:

    TALHOES   um talhão por linha: cultura, área, ano de plantio, fase de carga
    CURVAS    a curva de produtividade por idade, por cultura
    PRECOS    preço por unidade e custo por hectare, por estágio

A aba CURVAS existe porque a curva é a peça que o motor se recusa a inventar.
Ela precisa entrar por algum lugar, e o lugar é aqui — junto da coluna Fonte,
para o parecer poder dizer de onde veio.

Rode: python scripts/generate_ficha_talhoes.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = (
    Path(__file__).resolve().parent.parent
    / 'static' / 'templates' / 'ficha_talhoes_perene.xlsx'
)

IDADE_MAXIMA = 12

TALHOES = ('Identificação', 'Cultura', 'Área (ha)', 'Ano de plantio',
           'Fase de carga')
CURVAS = (['Cultura', 'Produtividade plena', 'Unidade', 'Bienalidade',
           'Ciclo (anos)', 'Fonte']
          + [f'Idade {i}' for i in range(1, IDADE_MAXIMA + 1)])
PRECOS = ('Cultura', 'Preço por unidade', 'Custo formação (R$/ha)',
          'Custo produção (R$/ha)', 'Custo reforma (R$/ha)',
          'Custo colheita por unidade')

_NEGRITO = Font(bold=True)
_FUNDO = PatternFill('solid', fgColor='D9E1F2')


def _cabecalho(ws, colunas, larguras: dict[int, int] | None = None) -> None:
    for indice, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=indice, value=titulo)
        celula.font = _NEGRITO
        celula.fill = _FUNDO
        celula.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[celula.column_letter].width = (
            (larguras or {}).get(indice, 18))
    ws.freeze_panes = 'A2'


def build() -> Workbook:
    wb = Workbook()

    talhoes = wb.active
    talhoes.title = 'TALHOES'
    _cabecalho(talhoes, TALHOES, {1: 22, 2: 16, 3: 12, 4: 16, 5: 16})
    fase = DataValidation(type='list', formula1='"alta,baixa"', allow_blank=True)
    talhoes.add_data_validation(fase)
    area = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                          allow_blank=True)
    talhoes.add_data_validation(area)
    for linha in range(2, 202):
        fase.add(talhoes.cell(row=linha, column=5))
        area.add(talhoes.cell(row=linha, column=3))

    curvas = wb.create_sheet('CURVAS')
    _cabecalho(curvas, CURVAS, {1: 16, 2: 18, 3: 14, 4: 13, 5: 13, 6: 30})

    precos = wb.create_sheet('PRECOS')
    _cabecalho(precos, PRECOS, {1: 16, 2: 18, 3: 20, 4: 20, 5: 20, 6: 22})

    notas = wb.create_sheet('LEIA-ME')
    for indice, texto in enumerate(_notas(), start=1):
        celula = notas.cell(row=indice, column=1, value=texto)
        if texto.endswith(':'):
            celula.font = _NEGRITO
    notas.column_dimensions['A'].width = 100
    return wb


def _notas() -> list[str]:
    return [
        'FICHA DE TALHÕES — LAVOURA PERENE',
        '',
        'TALHOES:',
        'Uma linha por talhão. O ano de plantio é o que define a idade, e a '
        'idade é o que define quanto o talhão produz.',
        'Fase de carga (alta/baixa) só vale para cultura com bienalidade, como '
        'o café: diz em que carga o talhão está na safra de referência.',
        'Talhões de um mesmo cafezal costumam estar em fases diferentes — é o '
        'que evita a lavoura inteira oscilar junto.',
        '',
        'CURVAS:',
        'Quanto um hectare produz em cada idade, como fração da produtividade '
        'plena. Idade 1 = primeiro ano após o plantio.',
        'Café: os primeiros anos são de formação (fator 0), depois a produção '
        'sobe até a plena.',
        'Cana: o decaimento da soqueira entra aqui — cada corte com fator menor '
        'que o anterior.',
        'Ciclo (anos): duração do ciclo do talhão até a reforma do canavial ou '
        'a recepa do cafezal. Em branco, o talhão não é reformado.',
        'Bienalidade: amplitude da alternância de carga (0,15 = 15% para mais '
        'no ano alto e para menos no baixo). Em branco ou zero, desliga.',
        'Fonte: de onde veio a curva. O parecer cita. Curva sem fonte é '
        'declaração do analista, e aparece como tal.',
        '',
        'PRECOS:',
        'Preço por unidade da curva — saca de 60 kg no café, tonelada na cana.',
        'O custo é por hectare e POR ESTÁGIO: o talhão em formação consome '
        'caixa por anos sem devolver nada, e um custo médio esconderia isso.',
        'Custo de colheita por unidade é opcional: colheita acompanha a '
        'produção, não a área.',
        '',
        'O QUE O SISTEMA NÃO INVENTA:',
        'Cultura sem curva não é projetada. Cultura sem preço não vira '
        'receita. Estágio sem custo declarado não recebe custo estimado.',
        'Em todos os casos a análise volta marcada como inválida, dizendo o '
        'que falta — em vez de devolver um número com aparência de cálculo.',
    ]


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    build().save(OUTPUT)
    print(f'Ficha gerada: {OUTPUT}')


if __name__ == '__main__':
    main()
