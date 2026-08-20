"""Gera a ficha de consolidação de rebanho por fazenda (.xlsx, sem macros).

Layout — é o contrato que services/importar_excel.py lê:
    Col B: rótulo feminino ou 'Fazenda' (nome da fazenda em col C)
    Col C: quantidade feminina
    Col E: rótulo masculino
    Col F: quantidade masculina
    Col H: total da fazenda (fórmula)
    Col J/K: 'Total Rebanho' — soma de todos os blocos (fórmula)

Um bloco por fazenda, BLOCOS ao todo. Sem VBA: a leitura dos PDFs é feita
pelo servidor (pdf_parsers.py), não por macro na máquina do analista.

Rode: python scripts/generate_ficha_consolidado.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = (
    Path(__file__).resolve().parent.parent
    / 'static' / 'templates' / 'ficha_consolidado_rebanho.xlsx'
)

BLOCOS = 10
LINHA_INICIAL = 3
ALTURA_BLOCO = 9

# (rótulo fêmea, rótulo macho) — as quatro faixas da ficha consolidada.
CATEGORIAS = [
    ('Bezerra', 'Bezerro'),
    ('Bezerra Desmama', 'Bezerro Desmama'),
    ('Novilha', 'Garrote'),
    ('Vaca', 'Boi Gordo'),
]


def _linha_do_bloco(indice: int) -> int:
    return LINHA_INICIAL + indice * ALTURA_BLOCO


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'CONSOLIDADO'

    negrito = Font(bold=True)
    fundo = PatternFill('solid', fgColor='D9E1F2')
    centro = Alignment(horizontal='center')

    qtd_dv = DataValidation(
        type='whole', operator='greaterThanOrEqual', formula1='0', allow_blank=True)
    ws.add_data_validation(qtd_dv)

    ws['B1'] = 'Ficha de consolidação de rebanho — Orkavyn Agro Intelligence'
    ws['B1'].font = Font(bold=True, size=12)

    for indice in range(BLOCOS):
        topo = _linha_do_bloco(indice)

        ws.cell(row=topo, column=2, value='Classificação Fêmea').font = negrito
        ws.cell(row=topo, column=5, value='Classificação Macho').font = negrito
        ws.cell(row=topo, column=8, value='Total Fazenda').font = negrito
        for coluna in (2, 5, 8):
            ws.cell(row=topo, column=coluna).fill = fundo

        ws.cell(row=topo + 1, column=2, value='Fazenda').font = negrito
        ws.cell(row=topo + 1, column=3, value=f'Fazenda {indice + 1}')
        ws.cell(row=topo + 1, column=5, value='Fazenda').font = negrito
        ws.cell(row=topo + 1, column=6, value=f'Fazenda {indice + 1}')

        for coluna, rotulo in ((2, 'Faixa'), (3, 'Qtd'), (5, 'Faixa'), (6, 'Qtd')):
            celula = ws.cell(row=topo + 2, column=coluna, value=rotulo)
            celula.font = negrito
            celula.alignment = centro

        for posicao, (femea, macho) in enumerate(CATEGORIAS):
            linha = topo + 3 + posicao
            ws.cell(row=linha, column=2, value=femea)
            ws.cell(row=linha, column=5, value=macho)
            qtd_dv.add(ws.cell(row=linha, column=3))
            qtd_dv.add(ws.cell(row=linha, column=6))

        rodape = topo + 3 + len(CATEGORIAS)
        primeira, ultima = topo + 3, rodape - 1
        ws.cell(row=rodape, column=2, value='Total').font = negrito
        ws.cell(row=rodape, column=5, value='Total').font = negrito
        ws.cell(row=rodape, column=3, value=f'=SUM(C{primeira}:C{ultima})').font = negrito
        ws.cell(row=rodape, column=6, value=f'=SUM(F{primeira}:F{ultima})').font = negrito
        ws.cell(row=rodape, column=8, value=f'=C{rodape}+F{rodape}').font = negrito

    _total_rebanho(ws, negrito, fundo)

    for coluna, largura in (('B', 20), ('C', 12), ('E', 20), ('F', 12),
                            ('H', 14), ('J', 20), ('K', 12)):
        ws.column_dimensions[coluna].width = largura
    return wb


def _total_rebanho(ws, negrito, fundo) -> None:
    """Bloco J/K: soma de cada categoria em todos os blocos de fazenda."""
    ws.cell(row=LINHA_INICIAL, column=10, value='Total Rebanho').font = negrito
    ws.cell(row=LINHA_INICIAL, column=10).fill = fundo
    ws.cell(row=LINHA_INICIAL + 1, column=10, value='Faixa').font = negrito
    ws.cell(row=LINHA_INICIAL + 1, column=11, value='Quantidade').font = negrito

    rotulos = [f for f, _ in CATEGORIAS] + [m for _, m in CATEGORIAS]
    for posicao, rotulo in enumerate(rotulos):
        coluna = 'C' if posicao < len(CATEGORIAS) else 'F'
        deslocamento = posicao % len(CATEGORIAS)
        celulas = '+'.join(
            f'{coluna}{_linha_do_bloco(i) + 3 + deslocamento}' for i in range(BLOCOS))
        linha = LINHA_INICIAL + 2 + posicao
        ws.cell(row=linha, column=10, value=rotulo)
        ws.cell(row=linha, column=11, value=f'={celulas}')

    linha_total = LINHA_INICIAL + 2 + len(rotulos)
    ws.cell(row=linha_total, column=10, value='Total').font = negrito
    ws.cell(
        row=linha_total, column=11,
        value=f'=SUM(K{LINHA_INICIAL + 2}:K{linha_total - 1})').font = negrito


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    build().save(OUTPUT)
    print(f'Ficha gerada: {OUTPUT}')


if __name__ == '__main__':
    main()
