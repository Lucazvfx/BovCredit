"""A ficha que distribuímos é nossa, e não carrega macro nem dado de produtor.

O projeto versionava e servia por /api/ficha/download uma planilha .xlsm de
terceiro: autoria na aba CONSOLIDADO, 297 KB de projeto VBA e uma aba LOG com
o rebanho de uma propriedade identificável. Três problemas no mesmo arquivo —
procedência, superfície de macro na máquina do analista e dado real no git.

A ficha atual é gerada por scripts/generate_ficha_consolidado.py e a tabela de
classificação vive em data/mapeamento_classificacao.csv. Estes testes existem
para que o arquivo antigo não volte por descuido.
"""
import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

import database as db
from services.fichas_rebanho.mapping_loader import load_mapping
from services.importar_excel import parsear_ficha_excel

RAIZ = Path(__file__).resolve().parent.parent
FICHA = RAIZ / 'static' / 'templates' / 'ficha_consolidado_rebanho.xlsx'


def _login():
    db.init_db()
    email = 'ficha@example.com'
    u = db.buscar_usuario_email(email)
    if not u:
        db.criar_usuario(email, 'Ficha', 'senha123')
        u = db.buscar_usuario_email(email)
    from app import app
    app.config['TESTING'] = True
    c = app.test_client()
    with c.session_transaction() as s:
        s['_user_id'] = str(u['id'])
    return c


def test_o_repositorio_nao_versiona_planilha_com_macro():
    """Um .xlsm no repositório é VBA de alguém viajando no nosso pacote."""
    com_macro = [p for p in RAIZ.rglob('*.xlsm') if '.git' not in p.parts]
    assert com_macro == []


def test_a_ficha_distribuida_nao_tem_vba_nem_autoria_de_terceiro():
    conteudo = zipfile.ZipFile(FICHA).namelist()
    assert not [nome for nome in conteudo if 'vba' in nome.lower()]

    wb = openpyxl.load_workbook(FICHA)
    celulas = [
        c.value for linha in wb['CONSOLIDADO'].iter_rows() for c in linha
        if isinstance(c.value, str)
    ]
    assert not [v for v in celulas if 'desenvolvido por' in v.lower()]


def test_a_ficha_nasce_vazia_de_rebanho():
    """A planilha antiga vinha com o rebanho de um produtor real preenchido."""
    assert parsear_ficha_excel(FICHA) == []


def test_a_ficha_preenchida_volta_pelo_importador():
    """Trocar o arquivo não pode quebrar quem preenche e importa."""
    wb = openpyxl.load_workbook(FICHA)
    ws = wb['CONSOLIDADO']
    ws['C4'] = 'Fazenda de Teste'
    ws['C6'], ws['C7'], ws['C8'], ws['C9'] = 100, 80, 60, 200
    ws['F6'], ws['F7'], ws['F8'], ws['F9'] = 90, 70, 40, 20
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fazendas = parsear_ficha_excel(buffer)

    assert [f['fazenda'] for f in fazendas] == ['Fazenda de Teste']
    assert fazendas[0]['total'] == 660
    # Bezerra/bezerro de 0–12 meses entram divididos entre as duas faixas jovens.
    assert fazendas[0]['valores'] == [50, 45, 50, 45, 80, 70, 60, 40, 200, 20]


def test_a_rota_de_download_entrega_a_ficha_sem_macro():
    resposta = _login().get('/api/ficha/download')

    assert resposta.status_code == 200
    assert resposta.headers['Content-Disposition'].endswith('.xlsx')
    assert not [n for n in zipfile.ZipFile(io.BytesIO(resposta.data)).namelist()
                if 'vba' in n.lower()]


@pytest.mark.parametrize('estado,sexo,faixa,esperado', [
    ('MT_DECLARACAO', 'FEMEA', '13 A 24 MESES', 'Bezerra Desmama'),
    ('TO_DECLARACAO', 'MACHO', '25 A 36 MESES', 'Garrote'),
    ('RO_DECLARACAO', 'FEMEA', 'ACIMA DE 36 MESES', 'Vaca'),
])
def test_o_mapeamento_vem_do_csv_versionado(estado, sexo, faixa, esperado):
    """A tabela é de-para factual: faixa do órgão estadual → nome zootécnico."""
    regra = load_mapping().lookup(estado, sexo, faixa)

    assert regra is not None
    assert regra.classificacao == esperado
