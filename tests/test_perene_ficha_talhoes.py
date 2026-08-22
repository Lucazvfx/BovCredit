"""Ficha de talhões: o que o analista preenche vira análise sem retrabalho.

A ficha existe porque a curva de produtividade é a peça que o motor se recusa
a inventar — ela tem de entrar por algum lugar. E porque uma análise que só
roda por `curl` não é usável por quem faz o trabalho.

Ficha meio preenchida não é erro: é o caso comum. O parser devolve o que
falta, item a item, em vez de estourar.
"""
import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

import database as db
from services.perennial_engine import analisar_lavoura_perene, parsear_ficha_talhoes

FICHA = (Path(__file__).resolve().parent.parent
         / 'static' / 'templates' / 'ficha_talhoes_perene.xlsx')

CREDITO = {
    'credito_valor': 1_500_000, 'prazo_meses': 72, 'juros_aa': 0.105,
    'carencia_meses': 24, 'sistema_amortizacao': 'sac', 'periodicidade_meses': 12,
}


def _preencher(talhoes=(('A', 'Café', 40, 2021, 'alta'),),
               curvas=(('Café', 30, 'saca_60kg', 0.15, None, 'Fonte X',
                        0, 0, 0.4, 0.8, 1.0),),
               precos=(('Café', 1400, 14000, 9000, None, 120),)):
    wb = openpyxl.load_workbook(FICHA)
    for linha in talhoes:
        wb['TALHOES'].append(list(linha))
    for linha in curvas:
        wb['CURVAS'].append(list(linha))
    for linha in precos:
        wb['PRECOS'].append(list(linha))
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _login():
    db.init_db()
    email = 'talhoes@example.com'
    usuario = db.buscar_usuario_email(email)
    if not usuario:
        db.criar_usuario(email, 'Talhoes', 'senha123')
        usuario = db.buscar_usuario_email(email)
    from app import app
    app.config['TESTING'] = True
    cliente = app.test_client()
    with cliente.session_transaction() as sessao:
        sessao['_user_id'] = str(usuario['id'])
    return cliente


# ── A ficha em branco ───────────────────────────────────────────────────────

def test_a_ficha_distribuida_nao_tem_macro():
    assert not [nome for nome in zipfile.ZipFile(FICHA).namelist()
                if 'vba' in nome.lower()]


def test_a_ficha_em_branco_diz_o_que_falta_em_vez_de_estourar():
    r = parsear_ficha_talhoes(FICHA, ano_base=2026)

    assert r['completo'] is False
    assert any('Nenhum talhão' in item for item in r['faltando'])


def test_as_tres_abas_existem():
    wb = openpyxl.load_workbook(FICHA)
    assert {'TALHOES', 'CURVAS', 'PRECOS'} <= set(wb.sheetnames)


# ── Round-trip ──────────────────────────────────────────────────────────────

def test_a_ficha_preenchida_vira_payload_e_roda_a_analise():
    """O critério da fase: preencher e importar sem retrabalho."""
    r = parsear_ficha_talhoes(_preencher(), ano_base=2026)

    assert r['completo'] is True
    assert r['culturas'] == ('CAFE',)
    assert r['area_total_ha'] == 40.0

    analise = analisar_lavoura_perene(dict(r['payload'], anos=6, credito=CREDITO))
    assert analise['valido'] is True
    assert analise['credito']['analysis']['pior_periodo']['ano'] > 1


def test_acento_e_caixa_da_cultura_nao_quebram_o_cruzamento():
    """'Café' na aba TALHOES e 'CAFE' na CURVAS têm de ser a mesma cultura."""
    r = parsear_ficha_talhoes(
        _preencher(talhoes=(('A', 'café', 40, 2021, 'alta'),),
                   curvas=(('CAFÉ', 30, 'saca_60kg', 0.15, None, 'F', 0, 0, 0.4, 0.8, 1.0),),
                   precos=(('Cafe', 1400, 14000, 9000, None, 120),)),
        ano_base=2026)

    assert r['completo'] is True


def test_a_curva_da_planilha_chega_com_as_idades_certas():
    r = parsear_ficha_talhoes(_preencher(), ano_base=2026)

    assert r['payload']['curvas']['CAFE']['fatores'] == {
        1: 0.0, 2: 0.0, 3: 0.4, 4: 0.8, 5: 1.0}


# ── O que falta ─────────────────────────────────────────────────────────────

def test_cultura_sem_curva_e_apontada_pelo_nome():
    r = parsear_ficha_talhoes(
        _preencher(talhoes=(('A', 'Café', 40, 2021, 'alta'),
                            ('K', 'Cana', 100, 2024, None))),
        ano_base=2026)

    assert r['completo'] is False
    assert any('Sem curva de produtividade para CANA' in i for i in r['faltando'])
    assert any('Sem preço declarado para CANA' in i for i in r['faltando'])


def test_talhao_sem_area_e_recusado_com_o_numero_da_linha():
    r = parsear_ficha_talhoes(
        _preencher(talhoes=(('A', 'Café', None, 2021, 'alta'),)), ano_base=2026)

    assert any('linha 2' in item and 'área' in item for item in r['faltando'])


def test_fase_de_carga_invalida_vira_aviso_e_nao_valor_errado():
    r = parsear_ficha_talhoes(
        _preencher(talhoes=(('A', 'Café', 40, 2021, 'media'),)), ano_base=2026)

    assert any('alta ou baixa' in item for item in r['faltando'])
    assert r['payload']['talhoes'][0]['fase_bienal'] is None


def test_curva_sem_fonte_entra_como_declaracao_do_analista():
    """Quem lê o parecer precisa saber que a curva não tem lastro citável."""
    r = parsear_ficha_talhoes(
        _preencher(curvas=(('Café', 30, 'saca_60kg', 0.15, None, '',
                            0, 0, 0.4, 0.8, 1.0),)),
        ano_base=2026)

    assert any('sem fonte' in aviso.lower() for aviso in r['avisos'])


# ── As rotas ────────────────────────────────────────────────────────────────

def test_a_rota_entrega_a_ficha_em_branco():
    resposta = _login().get('/api/perene/ficha/download')

    assert resposta.status_code == 200
    assert resposta.headers['Content-Disposition'].endswith('.xlsx')


def test_a_rota_importa_a_ficha_preenchida():
    dados = {'ficha': (_preencher(), 'talhoes.xlsx'), 'ano_base': '2026'}

    resposta = _login().post('/api/perene/importar-ficha', data=dados,
                             content_type='multipart/form-data')

    assert resposta.status_code == 200
    corpo = resposta.get_json()
    assert corpo['completo'] is True
    assert corpo['payload']['ano_base'] == 2026


def test_a_rota_recusa_arquivo_que_nao_e_planilha():
    dados = {'ficha': (io.BytesIO(b'nao sou uma planilha'), 'ficha.xlsx')}

    resposta = _login().post('/api/perene/importar-ficha', data=dados,
                             content_type='multipart/form-data')

    assert resposta.status_code == 400
